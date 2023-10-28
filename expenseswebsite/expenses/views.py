from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.conf import settings

from userpreferences.models import UserPreference
from .models import Category, Expense
from json import loads
from datetime import date, timedelta, datetime
import calendar
from reportlab.lib.pagesizes import letter, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from csv import writer
import os


@login_required(login_url='/authentication/login')
def index(request):
    expenses = Expense.objects.filter(owner=request.user)
    paginator = Paginator(expenses, 12)
    page_number = request.GET.get('page')
    page_obj = Paginator.get_page(paginator, page_number)
    currency = UserPreference.objects.get(user=request.user).currency
    context = {
        'expenses': expenses,
        'page_obj': page_obj,
        'currency': currency
    }

    return render(request, 'expenses/index.html', context)


@login_required(login_url='/authentication/login')
def search_expenses(request):
    if request.method == 'POST':
        search_string = loads(request.body).get('searchText')

        expenses = Expense.objects.filter(
            amount__istartswith=search_string, owner=request.user) | Expense.objects.filter(
            date__istartswith=search_string, owner=request.user) | Expense.objects.filter(
            description__icontains=search_string, owner=request.user) | Expense.objects.filter(
            category__icontains=search_string, owner=request.user)

        data = expenses.values()

        return JsonResponse(list(data), safe=False)


@login_required(login_url='/authentication/login')
def add_expense(request):
    categories = Category.objects.all()
    context = {
        'categories': categories,
        'values': request.POST
    }

    if request.method == 'GET':
        return render(request, 'expenses/add-expense.html', context)

    elif request.method == 'POST':
        amount = request.POST['amount']

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'expenses/add-expense.html', context)

        category = request.POST['category']
        description = request.POST['description']
        date = request.POST['expense_date']

        if not description:
            messages.error(request, 'Description is required')
            return render(request, 'expenses/add-expense.html', context)

        if not date:
            messages.error(request, 'Date is required')
            return render(request, 'expenses/add-expense.html', context)

        Expense.objects.create(
            owner=request.user,
            amount=amount,
            date=date,
            category=category,
            description=description
        )

        messages.success(request, 'Expense saved successfully')

        return redirect('expenses')


@login_required(login_url='/authentication/login')
def edit_expense(request, id):
    categories = Category.objects.all()
    expense = Expense.objects.get(pk=id)
    context = {
        'expense': expense,
        'values': expense,
        'categories': categories
    }

    if request.method == 'GET':
        return render(request, 'expenses/edit-expense.html', context)

    elif request.method == 'POST':
        amount = request.POST['amount']

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'expenses/edit-expense.html', context)

        category = request.POST['category']
        description = request.POST['description']
        date = request.POST['expense_date']

        if not description:
            messages.error(request, 'Description is required')
            return render(request, 'expenses/edit-expense.html', context)

        if not date:
            messages.error(request, 'Date is required')
            return render(request, 'expenses/edit-expense.html', context)

        expense.owner = request.user
        expense.amount = amount
        expense.date = date
        expense.category = category
        expense.description = description

        expense.save()

        messages.success(request, 'Expense updated successfully')

        return redirect('expenses')


@login_required(login_url='/authentication/login')
def delete_expense(request, id):
    expense = Expense.objects.get(pk=id)
    expense.delete()

    messages.success(request, 'Expense removed successfully')

    return redirect('expenses')


@login_required(login_url='/authentication/login')
def expense_week_summary(request):
    todays_date = date.today()
    current_week_start = todays_date - timedelta(days=todays_date.weekday())
    current_week_end = current_week_start + timedelta(days=6)

    expenses = Expense.objects.filter(
        owner=request.user,
        date__gte=current_week_start,
        date__lte=current_week_end
    )

    days_of_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    final_rep = {day: 0 for day in days_of_week}

    for expense in expenses:
        day_of_week = expense.date.strftime('%a')
        final_rep[day_of_week] += expense.amount

    return JsonResponse({'expense_week_data': final_rep}, safe=False)


@login_required(login_url='/authentication/login')
def expense_month_summary(request):
    todays_date = date.today()
    first_day_of_current_month = todays_date.replace(day=1)
    last_day_of_current_month = (todays_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    expenses = Expense.objects.filter(
        owner=request.user,
        date__gte=first_day_of_current_month,
        date__lte=last_day_of_current_month
    )

    num_days = last_day_of_current_month.day
    final_rep = {str(i): 0 for i in range(1, num_days + 1)}

    for expense in expenses:
        day = str(expense.date.day)
        final_rep[day] += expense.amount

    return JsonResponse({'expense_month_data': final_rep}, safe=False)


@login_required(login_url='/authentication/login')
def expense_year_summary(request):
    todays_date = date.today()
    first_day_of_current_year = todays_date.replace(month=1, day=1)
    last_day_of_current_year = todays_date.replace(month=12, day=31)

    expenses = Expense.objects.filter(
        owner=request.user,
        date__gte=first_day_of_current_year,
        date__lte=last_day_of_current_year
    )

    month_names = [calendar.month_abbr[i] for i in range(1, 13)]
    final_rep = {month: 0 for month in month_names}

    for expense in expenses:
        month_name = calendar.month_abbr[expense.date.month]
        final_rep[month_name] += expense.amount

    return JsonResponse({'expense_year_data': final_rep}, safe=False)


@login_required(login_url='/authentication/login')
def expense_card_summary(request):
    def get_value(expenses):
        count = 0
        amount = 0

        for expense in expenses:
            amount += expense.amount
            count += 1

        return {'count': count, 'amount': amount}

    todays_date = date.today()

    current_week_start = todays_date - timedelta(days=todays_date.weekday())
    current_week_end = current_week_start + timedelta(days=6)

    first_day_of_current_month = todays_date.replace(day=1)
    last_day_of_current_month = (todays_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    first_day_of_current_year = todays_date.replace(month=1, day=1)
    last_day_of_current_year = todays_date.replace(month=12, day=31)

    all_expenses = Expense.objects.filter(
        owner=request.user,
        date__gte=first_day_of_current_year,
        date__lte=last_day_of_current_year
    )

    today_expenses = all_expenses.filter(
        date=todays_date
    )

    week_expenses = all_expenses.filter(
        date__gte=current_week_start,
        date__lte=current_week_end
    )

    month_expenses = all_expenses.filter(
        date__gte=first_day_of_current_month,
        date__lte=last_day_of_current_month
    )

    final_rep = {
        'today': get_value(today_expenses),
        'week': get_value(week_expenses),
        'month': get_value(month_expenses),
        'year': get_value(all_expenses),
    }

    return JsonResponse({'expense_card_data': final_rep}, safe=False)


@login_required(login_url='/authentication/login')
def expense_stats(request):
    return render(request, 'expenses/expense-stats.html')


@login_required(login_url='/authentication/login')
def export_expenses_pdf(request):
    def pdf_page_template(canvas, doc):
        canvas.saveState()
        canvas.setFont('Times-Bold', 12)
        canvas.drawString(1 * inch, 0.75 * inch, 'YourExpenseManager')
        canvas.restoreState()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={request.user.username}_expenses_{str(datetime.now())}.pdf'

    doc = SimpleDocTemplate(response, pagesize=portrait(letter))
    story = []

    logo = os.path.join(settings.BASE_DIR, 'expenseswebsite', 'static', 'img', 'black_white_bg.png')
    im = Image(logo, 1 * inch, 1 * inch)
    story.append(im)

    styles = getSampleStyleSheet()
    style = styles["Heading1"]
    ptext = '<font size=16>Expense Report</font>'
    story.append(Paragraph(ptext, style))
    story.append(Spacer(1, 0.5 * inch))

    data = [['Amount', 'Date', 'Category', 'Description']]
    expenses = Expense.objects.filter(owner=request.user)
    for expense in expenses:
        data.append([expense.amount, expense.date, expense.category, expense.description])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(table)
    story.append(PageBreak())

    doc.build(story, onFirstPage=pdf_page_template, onLaterPages=pdf_page_template)

    return response


@login_required(login_url='/authentication/login')
def export_expenses_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={request.user.username}_expenses_{str(datetime.now())}.xlsx'

    wb = Workbook()
    ws = wb.create_sheet('Expenses', 0)

    header_font_style = Font(
        size=14,
        bold=True
    )

    header_alignment = Alignment(horizontal='center', vertical='center')
    main_alignment = Alignment(horizontal='left', vertical='center')

    column_width = 20
    columns = ['A', 'B', 'C', 'D']

    for column in columns:
        ws.column_dimensions[column].width = column_width

    border = Border(
        left=Side(border_style="thin", color='00000000'),
        right=Side(border_style="thin", color='00000000'),
        top=Side(border_style="thin", color='00000000'),
        bottom=Side(border_style="thin", color='00000000')
    )

    ws.append(['Amount', 'Category', 'Description', 'Date'])

    for cell in ws[1]:
        cell.font = header_font_style
        cell.alignment = header_alignment
        cell.border = border

    rows = Expense.objects.filter(owner=request.user).values_list('amount', 'category', 'description', 'date')
    for row_num, row in enumerate(rows):
        ws.append(row)
        for cell in ws[2 + row_num]:
            cell.alignment = main_alignment
            cell.border = border

    wb.save(response)

    return response


@login_required(login_url='/authentication/login')
def export_expenses_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={request.user.username}_expenses_{str(datetime.now())}.csv'

    csv_writer = writer(response)
    csv_writer.writerow(['Amount', 'Category', 'Description', 'Date'])

    expenses = Expense.objects.filter(owner=request.user)

    for expense in expenses:
        csv_writer.writerow([expense.amount, expense.category, expense.description, expense.date])

    return response


@login_required(login_url='/authentication/login')
def export_expenses(request):
    return render(request, 'expenses/export-expenses.html')
