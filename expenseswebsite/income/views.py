from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse

from userpreferences.models import UserPreference
from .models import Source, Income
from json import loads
from datetime import date, timedelta, datetime
import calendar
from csv import writer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


@login_required(login_url='/authentication/login')
def index(request):
    income = Income.objects.filter(owner=request.user)
    paginator = Paginator(income, 12)
    page_number = request.GET.get('page')
    page_obj = Paginator.get_page(paginator, page_number)
    currency = UserPreference.objects.get(user=request.user).currency
    context = {
        'income': income,
        'page_obj': page_obj,
        'currency': currency
    }

    return render(request, 'income/index.html', context)


@login_required(login_url='/authentication/login')
def search_income(request):
    if request.method == 'POST':
        search_string = loads(request.body).get('searchText')

        income = Income.objects.filter(
            amount__istartswith=search_string, owner=request.user) | Income.objects.filter(
            date__istartswith=search_string, owner=request.user) | Income.objects.filter(
            description__icontains=search_string, owner=request.user) | Income.objects.filter(
            source__icontains=search_string, owner=request.user)

        data = income.values()

        return JsonResponse(list(data), safe=False)


@login_required(login_url='/authentication/login')
def add_income(request):
    sources = Source.objects.all()
    context = {
        'sources': sources,
        'values': request.POST
    }

    if request.method == 'GET':
        return render(request, 'income/add-income.html', context)

    elif request.method == 'POST':
        amount = request.POST['amount']

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'income/add-income.html', context)

        source = request.POST['source']
        description = request.POST['description']
        date = request.POST['income_date']

        if not description:
            messages.error(request, 'Description is required')
            return render(request, 'income/add-income.html', context)

        if not date:
            messages.error(request, 'Date is required')
            return render(request, 'income/add-income.html', context)

        Income.objects.create(
            owner=request.user,
            amount=amount,
            date=date,
            source=source,
            description=description
        )

        messages.success(request, 'Income saved successfully')

        return redirect('income')


@login_required(login_url='/authentication/login')
def edit_income(request, id):
    sources = Source.objects.all()
    income = Income.objects.get(pk=id)
    context = {
        'income': income,
        'values': income,
        'sources': sources
    }

    if request.method == 'GET':
        return render(request, 'income/edit-income.html', context)

    elif request.method == 'POST':
        amount = request.POST['amount']

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'income/edit-income.html', context)

        source = request.POST['source']
        description = request.POST['description']
        date = request.POST['income_date']

        if not description:
            messages.error(request, 'Description is required')
            return render(request, 'income/edit-income.html', context)

        if not date:
            messages.error(request, 'Date is required')
            return render(request, 'income/edit-income.html', context)

        income.owner = request.user
        income.amount = amount
        income.date = date
        income.source = source
        income.description = description

        income.save()

        messages.success(request, 'Income updated successfully')

        return redirect('income')


@login_required(login_url='/authentication/login')
def delete_income(request, id):
    income = Income.objects.get(pk=id)
    income.delete()

    messages.success(request, 'Income removed successfully')

    return redirect('income')


@login_required(login_url='/authentication/login')
def income_week_summary(request):
    todays_date = date.today()
    current_week_start = todays_date - timedelta(days=todays_date.weekday())
    current_week_end = current_week_start + timedelta(days=6)

    income = Income.objects.filter(
        owner=request.user,
        date__gte=current_week_start,
        date__lte=current_week_end
    )

    days_of_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    final_rep = {day: 0 for day in days_of_week}

    for item in income:
        day_of_week = item.date.strftime('%a')
        final_rep[day_of_week] += item.amount

    return JsonResponse({'income_week_data': final_rep}, safe=False)


@login_required(login_url='/authentication/login')
def income_month_summary(request):
    todays_date = date.today()
    first_day_of_current_month = todays_date.replace(day=1)
    last_day_of_current_month = (todays_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    income = Income.objects.filter(
        owner=request.user,
        date__gte=first_day_of_current_month,
        date__lte=last_day_of_current_month
    )

    num_days = last_day_of_current_month.day
    final_rep = {str(i): 0 for i in range(1, num_days + 1)}

    for item in income:
        day = str(item.date.day)
        final_rep[day] += item.amount

    return JsonResponse({'income_month_data': final_rep}, safe=False)


@login_required(login_url='/authentication/login')
def income_year_summary(request):
    todays_date = date.today()
    first_day_of_current_year = todays_date.replace(month=1, day=1)
    last_day_of_current_year = todays_date.replace(month=12, day=31)

    income = Income.objects.filter(
        owner=request.user,
        date__gte=first_day_of_current_year,
        date__lte=last_day_of_current_year
    )

    month_names = [calendar.month_abbr[i] for i in range(1, 13)]
    final_rep = {month: 0 for month in month_names}

    for item in income:
        month_name = calendar.month_abbr[item.date.month]
        final_rep[month_name] += item.amount

    return JsonResponse({'income_year_data': final_rep}, safe=False)


@login_required(login_url='/authentication/login')
def income_card_summary(request):
    def get_value(income):
        count = 0
        amount = 0

        for item in income:
            amount += item.amount
            count += 1

        return {'count': count, 'amount': amount}

    todays_date = date.today()

    current_week_start = todays_date - timedelta(days=todays_date.weekday())
    current_week_end = current_week_start + timedelta(days=6)

    first_day_of_current_month = todays_date.replace(day=1)
    last_day_of_current_month = (todays_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    first_day_of_current_year = todays_date.replace(month=1, day=1)
    last_day_of_current_year = todays_date.replace(month=12, day=31)

    all_income = Income.objects.filter(
        owner=request.user,
        date__gte=first_day_of_current_year,
        date__lte=last_day_of_current_year
    )

    today_income = all_income.filter(
        date=todays_date
    )

    week_income = all_income.filter(
        date__gte=current_week_start,
        date__lte=current_week_end
    )

    month_income = all_income.filter(
        date__gte=first_day_of_current_month,
        date__lte=last_day_of_current_month
    )

    final_rep = {
        'today': get_value(today_income),
        'week': get_value(week_income),
        'month': get_value(month_income),
        'year': get_value(all_income),
    }

    return JsonResponse({'income_card_data': final_rep}, safe=False)


@login_required(login_url='/authentication/login')
def income_stats(request):
    return render(request, 'income/income-stats.html')


@login_required(login_url='/authentication/login')
def export_income_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={request.user.username}_income_{str(datetime.now())}.xlsx'

    wb = Workbook()
    ws = wb.create_sheet('Income', 0)

    header_font_style = Font(
        size=14,
        bold=True
    )

    header_alignment = Alignment(horizontal='center', vertical='center')
    main_alignment = Alignment(horizontal='left', vertical='center')

    column_width = 20
    columns = ['A', 'B', 'C', 'D']

    for column in columns:
        ws.column_dimensions[column].width = column_width

    border = Border(
        left=Side(border_style="thin", color='00000000'),
        right=Side(border_style="thin", color='00000000'),
        top=Side(border_style="thin", color='00000000'),
        bottom=Side(border_style="thin", color='00000000')
    )

    ws.append(['Amount', 'Source', 'Description', 'Date'])

    for cell in ws[1]:
        cell.font = header_font_style
        cell.alignment = header_alignment
        cell.border = border

    rows = Income.objects.filter(owner=request.user).values_list('amount', 'source', 'description', 'date')
    for row_num, row in enumerate(rows):
        ws.append(row)
        for cell in ws[2 + row_num]:
            cell.alignment = main_alignment
            cell.border = border

    wb.save(response)

    return response


@login_required(login_url='/authentication/login')
def export_income_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={request.user.username}_income_{str(datetime.now())}.csv'

    csv_writer = writer(response)
    csv_writer.writerow(['Amount', 'Source', 'Description', 'Date'])

    income = Income.objects.filter(owner=request.user)

    for item in income:
        csv_writer.writerow([item.amount, item.source, item.description, item.date])

    return response


@login_required(login_url='/authentication/login')
def export_income(request):
    return render(request, 'income/export-income.html')
